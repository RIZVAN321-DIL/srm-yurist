from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_session
from app.repositories.client_repo import ClientRepository
from app.api.schemas.client import ClientCreate, ClientUpdate
from app.models.client import Client
from app.core.auth_middleware import get_current_user
import openpyxl, io
from fastapi.responses import StreamingResponse

router = APIRouter(prefix="/api", tags=["clients"])

@router.get("/clients")
async def get_clients(request: Request, status=None, search=None, skip=0, limit=50, session=Depends(get_session), user=Depends(get_current_user)):
    clients = await ClientRepository(session).get_all(status=status, search=search, skip=skip, limit=limit)
    return [{"id": c.id, "full_name": c.full_name, "phone": c.phone, "email": c.email, "status": c.status, "tags": c.tags, "notes": c.notes, "access_code": c.access_code} for c in clients]

@router.get("/clients/{client_id}")
async def get_client(client_id: int, session=Depends(get_session), user=Depends(get_current_user)):
    c = await ClientRepository(session).get_by_id(client_id)
    if not c: raise HTTPException(404)
    return {"id": c.id, "full_name": c.full_name, "phone": c.phone, "email": c.email, "status": c.status, "tags": c.tags, "notes": c.notes, "access_code": c.access_code}

@router.post("/clients")
async def create_client(data: ClientCreate, session=Depends(get_session), user=Depends(get_current_user)):
    c = Client(full_name=data.full_name, phone=data.phone, email=data.email, status=data.status, tags=data.tags, notes=data.notes)
    result = await ClientRepository(session).create(c); await session.commit(); return {"ok": True, "id": result.id}

@router.put("/clients/{client_id}")
async def update_client(client_id: int, data: ClientUpdate, session=Depends(get_session), user=Depends(get_current_user)):
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    ok = await ClientRepository(session).update(client_id, **updates)
    if not ok: raise HTTPException(404)
    await session.commit(); return {"ok": True}

@router.delete("/clients/{client_id}")
async def delete_client(client_id: int, session=Depends(get_session), user=Depends(get_current_user)):
    ok = await ClientRepository(session).delete(client_id)
    if not ok: raise HTTPException(404)
    await session.commit(); return {"ok": True}

@router.get("/clients/export/excel")
async def export_clients(session=Depends(get_session), user=Depends(get_current_user)):
    clients = await ClientRepository(session).get_all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Клиенты"
    ws.append(["ID", "ФИО", "Телефон", "Email", "Статус", "Теги", "Заметки"])
    for c in clients: ws.append([c.id, c.full_name, c.phone, c.email, c.status, c.tags, c.notes])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=clients.xlsx"})

@router.post("/clients/import/excel")
async def import_clients(file: UploadFile, session=Depends(get_session), user=Depends(get_current_user)):
    content = await file.read(); wb = openpyxl.load_workbook(io.BytesIO(content)); ws = wb.active; count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            c = Client(full_name=str(row[1]), phone=str(row[2] or ""), email=str(row[3] or ""), status=str(row[4] or "active"))
            await ClientRepository(session).create(c); count += 1
    await session.commit(); return {"ok": True, "imported": count}
